import argparse
import json
import os
from openpyxl.utils import get_column_letter
import pandas as pd


def process_trafa_data(): 
    # Specify the folder path where the JSON files are located
    folder_path = 'data/trafa/downloads'

    # Get the file named 'trafa-data.json'
    file = 'trafa-data.json'
    
    # Read the JSON file into a DataFrame
    with open(os.path.join(folder_path, file)) as f:
        data = json.load(f)
        print(f"Processing {file}...")
    
    df = pd.DataFrame()
    # Iterate over each row
    for row in data:
        # Extract the data under the "Cell" key
        cell_data = row['municipalities']
          
        # Iterate over each cell
        for cell in cell_data:
            # Add the cell data to the DataFrame
            percentage = float(cell['percentageElectricCars'])
        
            df.loc[cell['name'], cell['year']] = percentage

    df = df.reset_index()
    df.columns = ['Kommun'] + list(df.columns[1:])
    for index, row in df.iterrows():       
        # Excel is 1-based and the first row is the header row so we need to add 2 to the index to get the correct column
        if index == 0:
            length = len(df.columns) + 1
        else:
            length = len(df.columns) 
        df.loc[index, 'Trend'] = f'=LINEST(C{index + 2}:{get_column_letter(len(df.columns))}{index + 2}, $C$1:${get_column_letter(length)}$1, TRUE)'  # type: ignore

    # Write the DataFrame to an Excel file
    output_file = 'data/output/trafa-output.xlsx'

    df.to_excel(output_file)
    print(f"Data successfully transformed and saved to {output_file}") 


def compare_trafa_data():
    # Read the Excel file into a DataFrame
    df_trafa = pd.read_excel('data/output/trafa-output.xlsx')
    # df_trafa = df_trafa.drop([0])
    
    # df_trafa = df_trafa.rename(columns={'Unnamed: 0': 'Kommun'}) 
    df_trafa = df_trafa.set_index('Kommun')
    
    # drop the last comlumn of the DataFrame
    df_trafa = df_trafa.drop([df_trafa.columns[0],df_trafa.columns[-1],df_trafa.columns[-2]], axis=1)
 



    # Read the Excel file into a DataFrame
    df_cars = pd.read_excel('data/solutions/cars/sources/kpi1_calculations.xlsx')
        
    df_cars = df_cars.drop([0])
    
    # drop the last 2 columns of the DataFrame
    df_cars = df_cars.drop(df_cars.columns[-2:], axis=1)
    
    # use the first row as the column names 
    df_cars.columns = ["Kommun"] + list(df_cars.iloc[0].values[1:9].astype(int))
    df_cars = df_cars.drop([1])
    
    df_cars = df_cars.set_index('Kommun')
    # drop the last row of the DataFrame
    df_cars = df_cars.drop(df_cars.index[-1])
    
    
    # Multiply the values of the DataFrame by 100 to get the percentage and round to 2 decimal places
    df_cars = df_cars * 100
    df_cars = df_cars.round(1)
    
    
    # order the index of the DataFrame 
    df_cars = df_cars.sort_index()
    df_trafa = df_trafa.sort_index()

    
    # compare the columns of the two DataFrames
    if df_cars.columns.equals(df_trafa.columns):
        print("The columns of the two DataFrames are equal")
    else:
        print("The columns of the two DataFrames are not equal")
        
    # compare the indices of the two DataFrames
    if df_cars.index.equals(df_trafa.index):
        print("The indices of the two DataFrames are equal")
    else:
        print("The indices of the two DataFrames are not equal")
        
    # compare the values of the two DataFrames
    if df_cars.equals(df_trafa):
        print("The two DataFrames are equal")
    else:
        print("The two DataFrames are not equal")
        
        
  #  Find the differences between the two DataFrames and store them in a list
    comparisson = []
    for index, row in df_cars.iterrows():
        for column in df_cars.columns:
            kpi1 = row[column]
            if index in df_trafa.index and column in df_trafa.columns:
                tra = df_trafa.loc[index, column] # type: ignore
                if kpi1 != tra:
                    comparisson.append([index, column, kpi1, tra])
            else:
                print(f"Index {index} or column {column} not found in df_trafa")
                
    # Create a DataFrame from the comparison list
    df_comparisson = pd.DataFrame(comparisson, columns=['Kommun', 'År', 'KPI1', 'Trafa-API'])
    
    # Write the DataFrame to an Excel file
    output_file = 'data/output/trafa-comparison.xlsx'
    df_comparisson.to_excel(output_file, index=False)
                        

    print("Data successfully compared and saved to data/output/trafa-comparison.xlsx")
    
    

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--function", help="The function to run")
    parser.add_argument("--json-data", help="The JSON data to process")
    args = parser.parse_args()
    
    if args.function == "process_trafa_data":
        process_trafa_data()
    if args.function == "compare_trafa_data":
        compare_trafa_data()

